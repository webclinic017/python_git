# from struct import pack, unpack
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time

# start = time.time()

def twos_comp(val, bits):
    """compute the 2's complement of int value val"""
    if (val & (1 << (bits - 1))) != 0: # if sign bit is set e.g., 8bit: 128-255
        val = val - (1 << bits)        # compute negative value
    return val                         # return positive value as is

gen_file = "C:\\Users\\Brian.tsai\\Desktop\\Python\\STDF\\STDF.xlsx"
path_dir = 'main_Lot_1_Wafer_1_Oct_13_09h33m41s_STDF'
size = os.path.getsize(path_dir)
# print(size)
print('%s = %d bytes' % (path_dir, size))
data = []
PRR_MAP = {}
f = open("C:\\Users\\Brian.tsai\\Desktop\\Python\\STDF\\main_Lot_1_Wafer_1_Oct_13_09h33m41s_STDF", mode = 'rb')
for i in range(0,size):
    data += f.read(1)
# print(data)    
for i in range(0,size):
    if (data[i] == 5 and data[i+1] == 20): 
        coordinate = ""
        # print(bin(data[i+11]))
        coord_x_big = bin(data[i+12])[2:]  #找出十位數, [2:]表示不要取0b, 預設bin強制轉型的string會帶有"0b"
        # print(coord_x_big)
        coord_x_small = bin(data[i+11])[2:] #找出個位數字
        # print(coord_x_small)
        coord_x = coord_x_big + coord_x_small # coord_x_big 與 coord_x_small 皆為字串可以直接用"+"將字串合併
        # print(coord_x)
        # print(int(coord_x,2))
        # print(int(coord_x))
        # print(int(coord_x,4))
        coord_x =  twos_comp(int(coord_x,2), 16) #使用twos_comp做2補數, (coord_x, 2)代表將字串轉為10進位, int((coord_x, 2))的int是強制轉型, 由於twos_comp這個方法需要丟int的型態進去
        # print(bin(data[i+12]<<8))
        # print(int(coord_x))

        coord_y_big = bin(data[i+14])[2:]
        coord_y_small = bin(data[i+13])[2:]
        coord_y = coord_y_big + coord_y_small
        coord_y = twos_comp(int(coord_y,2), 16)
        # coord_x = str(twos_comp(int((str(bin(data[i+12]))[2:] + str(bin(data[i+11]))[2:]),2),16))
        # coord_y = str(twos_comp(int((str(bin(data[i+14]))[2:] + str(bin(data[i+13]))[2:]),2),16))
        # print("coord_x = " +  str(coord_x))
        # print("coord_y = " +  str(coord_y))
        coordinate = str(coord_x) + "," + str(coord_y) 
        if (data[i+19] == 1):
            part_ID = chr(data[i+20])
            # print("part_ID = " + part_ID)
            PRR_MAP[part_ID] = coordinate
            
        if (data[i+19] == 2):
            part_ID = chr(data[i+20]) + chr(data[i+21])
            # print("part_ID = " + part_ID)
            PRR_MAP[part_ID] = coordinate
print(PRR_MAP)            
f.close()            
# for key in PRR_MAP:
#     print("part_ID: " + key)
#     print("X, Y: " + PRR_MAP[key])

fill_red = PatternFill('solid', fgColor='FF0000')
fill_green = PatternFill('solid', fgColor='00FF00')

wb_gen = Workbook()
ws_gen = wb_gen["Sheet"]
ws_gen.title = "STDF"    

X_mapping = {}
Y_mapping = {}
col_num = 2
for i in range (-7, 8, 1):
    ws_gen.cell(row=1, column=col_num, value=i)
    ws_gen.cell(row=1, column=col_num).fill = fill_green
    X_mapping[i] = col_num
    col_num = col_num + 1

row_num = 2
for i in range (4, -5, -1):
    ws_gen.cell(row=row_num, column=1, value=i)
    # ws_gen.cell(row=row_num, column=1).fill = fill_green
    Y_mapping[i] = row_num
    row_num = row_num + 1
print(X_mapping)
print(Y_mapping)

for key in PRR_MAP.keys():
    content = PRR_MAP[key]
    # print(content)
    content = content.split(",")
    print(content)
    ws_gen.cell(row=Y_mapping[int(content[1])], column=X_mapping[int(content[0])], value=key)
    # ws_gen.cell(row=Y_mapping[int(content[1])], column=X_mapping[int(content[0])]).fill = fill_red

wb_gen.save(gen_file)

# end = time.time()
# print("Program time: " + str(end-start) + "(s)")
