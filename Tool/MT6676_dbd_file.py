from tool_function import gen_pathList
import os
import re
from openpyxl import load_workbook

try:
    os.remove("D:\\Project\\MT6676\\javaCode.txt")
except OSError as e:
    print(e)
else:
    print("================================File is deleted successfully================================")

path_list = gen_pathList("D:\\Project\\MT6676","xlsx")
# print(path_list)
file_java = open("D:\\Project\\MT6676\\javaCode.txt",'w') 
for file in path_list:
    wb = load_workbook(file)
    # print(wb.sheetnames)
    for sheet in wb.sheetnames: 
        if sheet == 'DUT':
            signal_list = []
            pin_list = []
            merge_list = []
            pattern_D = r'\w*D\w*'
            pattern_DGND = r'\w*DGND\w*'
            pattern_AVI64_d = r'\w*AVI\d*\D\d'
            pattern_AVI64_dd = r'\w*AVI\d*\D\d\d'   
            ws_DUT = wb['DUT']

            for i in range(3,ws_DUT.max_row-1, 1):
                Ce = (ws_DUT.cell(row = i, column=2))
                signal_names = Ce.value
                signal_list.append(signal_names)
            # print(signal_list)  ## signal names

            for k in range(3,ws_DUT.max_row-1, 1):
                pin_list_tmp = []
                for i in range(3,ws_DUT.max_column+1, 2):           
                    Ce = (ws_DUT.cell(row = k, column=i))
                    pin = Ce.value
                    pin_list_tmp.append(pin)
                pin_list.append(pin_list_tmp)
                pin_list_tmp = []    
            # print(pin_list)
            begin = 0
            for list_tmp in pin_list:
                begin_index = 0
                for item in list_tmp:
                    if re.search(pattern_D, item) != None:  
                        item = item[1:]
                        pin_list[begin][begin_index] = item
                        begin_index = begin_index + 1
                        # print(begin_index)
                    elif re.search(pattern_AVI64_dd, item) != None:  
                        # print(item)       
                        item = item[5:8] + item[9:11]
                        pin_list[begin][begin_index] = item
                        begin_index = begin_index + 1
                        # print(item)
                    elif re.search(pattern_AVI64_d, item) != None:    
                        # print(item)
                        item = item[5:8] + str(0) + item[9]
                        # print(item)
                        pin_list[begin][begin_index] = item
                        begin_index = begin_index + 1
                        # print(begin_index)            
                begin = begin + 1                
            # print(pin_list)      
            for i in range(4,ws_DUT.max_column+1, 2):
                Ce = (ws_DUT.cell(row = 7, column=i))
                # print(Ce.value)      
                merge = Ce.value.split("\n")
                merge_list.append(merge)
            # print(merge_list)
            begin = 0
            for list in merge_list:
                begin_index = 0
                for item in list:
                    if re.search(pattern_AVI64_dd, item) != None:  
                        # print(item)       
                        item = item[5:8] + item[9:11]
                        merge_list[begin][begin_index] = item
                        begin_index = begin_index + 1
                    elif re.search(pattern_AVI64_d, item) != None:    
                        # print(item)
                        item = item[5:8] + str(0) + item[9]
                        # print(item)
                        merge_list[begin][begin_index] = item
                        begin_index = begin_index + 1    
                begin = begin + 1   
            # print(merge_list)                       
            begin = 0
            begin_index = 0
            iSite = 1             
            for signal in signal_list:
                # print(signal)
                file_java.write("signal " + signal + "{\n")
                # time.sleep(0.001)
                for pin in pin_list[begin]:
                    # print(pin)
                    if iSite == 16:
                        # print(begin)
                        file_java.write("Site " + str(iSite) + "{ pogo = " + pin+ "; }\n")
                        file_java.write("}\n")    
                        iSite = 1
                    else:    
                        file_java.write("Site " + str(iSite) + "{ pogo = " + pin+ "; }\n")
                        iSite = iSite + 1
                begin = begin + 1        
            file_java.write("signal " + "AVI64_LED" + "{\n")
            iSite = 1       
            for pin in merge_list:
                if iSite == 16:
                    # print(begin)
                    file_java.write("Site " + str(iSite) + "{ pogo = " + pin[0] + "|" + pin[1] + "|" + pin[2] + "; }\n")
                    file_java.write("}\n")    
                    iSite = 1
                else:
                    file_java.write("Site " + str(iSite) + "{ pogo = " + pin[0] + "|" + pin[1] + "|" + pin[2] + "; }\n")   
                    iSite = iSite + 1
            # file_java.close

        elif sheet == 'OSC':    
            pin_list = []
            ws_OSC = wb['OSC']
            begin = 0
            pattern_D = r'\w*D\w*'
            for i in range(2, ws_OSC.max_row + 1, 1):
                Ce = (ws_OSC.cell(row = i, column=2))
                pin = Ce.value
                pin_list.append(pin)
            # print(pin_list)
            for pin in pin_list:
                pin = pin[1:]
                pin_list[begin] = pin
                begin = begin + 1
            # print(pin_list)
            file_java.write("signal " + "OSC" + "{\n")
            iSite = 1       
            for pin in pin_list:
                if iSite == 16:
                    # print(begin)
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")
                    file_java.write("}\n")    
                    iSite = 1
                else:
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")   
                    iSite = iSite + 1

        elif sheet == 'BUFFER':
            pin_list = []
            AVI64_list = []
            begin = 0
            ws_BUFFER = wb['BUFFER']
            pattern_AVI64_d = r'\w*AVI\d*\D\d'
            pattern_AVI64_dd = r'\w*AVI\d*\D\d\d'   
            # print(ws_OSC.max_row)
            for i in range(1, 32 + 1, 2):
                Ce = (ws_BUFFER.cell(row = i, column=2))
                pin = Ce.value
                pin_list.append(pin)
            # print(pin_list)
            for pin in pin_list:
                pin = pin[1:]
                pin_list[begin] = pin
                begin = begin + 1    
            # print(pin_list)
            for i in range(2, 32 + 1, 2):
                Ce = (ws_BUFFER.cell(row = i, column=2))
                pin = Ce.value
                AVI64_list.append(pin)
            # print(AVI64_list)
            begin = 0
            for AVI64 in AVI64_list:
                if re.search(pattern_AVI64_dd, AVI64) != None: 
                    AVI64 = AVI64[5:8] + AVI64[9:11]
                    AVI64_list[begin] = AVI64
                    begin = begin + 1
                elif re.search(pattern_AVI64_d, AVI64) != None:  
                    AVI64 = AVI64[5:8] + str(0) + AVI64[9]
                    # print(item)
                    AVI64_list[begin] = AVI64
                    begin = begin + 1    
            # print(AVI64_list)      
            file_java.write("signal " + "BUFFER_D" + "{\n")
            iSite = 1  
            for pin in pin_list:
                if iSite == 16:
                    # print(begin)
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")
                    file_java.write("}\n")    
                    iSite = 1
                else:
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")   
                    iSite = iSite + 1
            file_java.write("signal " + "BUFFER_AVI64" + "{\n")
            iSite = 1  
            for pin in AVI64_list:
                if iSite == 16:
                    # print(begin)
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")
                    file_java.write("}\n")    
                    iSite = 1
                else:
                    file_java.write("Site " + str(iSite) + "{ pogo = "  + pin  + "; }\n")   
                    iSite = iSite + 1    
        elif sheet == 'UTILITY':
            pin_list = []
            AVI64_list = []
            begin = 0
            ws_UTILITY = wb['UTILITY']
            K1_list = []
            K2_list = []
            K3_list = []
            K4_list = []
            K5_list = []
            K6_list = []
            K7_list = []
            K8_list = []
            K9_list = []
            K10_list = []
            K11_list = []
            K12_list = []
            K13_list = []
            K14_list = []
            UTILITY_list = []
            # print(ws_OSC.max_row)
            for k in range (2, 16, 1):
                for i in range(k, 226, 14):
                    Ce = (ws_UTILITY.cell(row = i, column=2))
                    pin = Ce.value
                    if k == 2:
                        K1_list.append(pin)
                    elif k == 3:
                        K2_list.append(pin)
                    elif k == 4:
                        K3_list.append(pin)
                    elif k == 5:
                        K4_list.append(pin)
                    elif k == 6:
                        K5_list.append(pin)
                    elif k == 7:
                        K6_list.append(pin)
                    elif k == 8:
                        K7_list.append(pin)
                    elif k == 9:
                        K8_list.append(pin)
                    elif k == 10:
                        K9_list.append(pin)
                    elif k == 11:
                        K10_list.append(pin)
                    elif k == 12:
                        K11_list.append(pin)
                    elif k == 13:
                        K12_list.append(pin)
                    elif k == 14:
                        K13_list.append(pin)
                    elif k == 15:
                        K14_list.append(pin)
            UTILITY_list.append(K1_list)     
            UTILITY_list.append(K2_list) 
            UTILITY_list.append(K3_list) 
            UTILITY_list.append(K4_list) 
            UTILITY_list.append(K5_list) 
            UTILITY_list.append(K6_list) 
            UTILITY_list.append(K7_list) 
            UTILITY_list.append(K8_list) 
            UTILITY_list.append(K9_list) 
            UTILITY_list.append(K10_list) 
            UTILITY_list.append(K11_list) 
            UTILITY_list.append(K12_list) 
            UTILITY_list.append(K13_list) 
            UTILITY_list.append(K14_list)         
            # print(UTILITY_list)
            begin = 0
            for relay in UTILITY_list:
                begin_index = 0
                for pin in relay:
                    pin = str(int(pin[2] + pin[7:9]) + 1)
                    UTILITY_list[begin][begin_index] = pin
                    begin_index = begin_index + 1
                begin = begin + 1    
            # print(UTILITY_list)   
            iSite = 1             
            relay = 1
            for UTILITY in UTILITY_list:
                # print(signal)
                file_java.write("signal K" + str(relay) + "{\n")
                for pin in UTILITY:
                    # print(pin)
                    if iSite == 16:
                        # print(begin)
                        file_java.write("Site " + str(iSite) + "{ pogo = " + pin+ "; }\n")
                        file_java.write("}\n")    
                        iSite = 1
                    else:    
                        file_java.write("Site " + str(iSite) + "{ pogo = " + pin+ "; }\n")
                        iSite = iSite + 1
                relay = relay + 1        
            KB6_list = []
            KB7_list = [] 
            KB8_list = [] 
            KB24_list = [] 
            KB34_list = [] 
            KB35_list = [] 
            KB36_list = [] 
            KB37_list = [] 
            KB38_list = []    
            KB_list = []         
            for k in range (226, 226 + 9, 1):
                for i in range(k, 244, 9):
                    Ce = (ws_UTILITY.cell(row = i, column=2))
                    pin = Ce.value
                    if k == 226:
                        KB6_list.append(pin)
                    elif k == 227:
                        KB7_list.append(pin)  
                    elif k == 228:
                        KB8_list.append(pin)   
                    elif k == 229:
                        KB24_list.append(pin)
                    elif k == 230:
                        KB34_list.append(pin)
                    elif k == 231:
                        KB35_list.append(pin)
                    elif k == 232:
                        KB36_list.append(pin)
                    elif k == 233:
                        KB37_list.append(pin)
                    elif k == 234:
                        KB38_list.append(pin)                           
            KB_list.append(KB6_list)
            KB_list.append(KB7_list)
            KB_list.append(KB8_list)
            KB_list.append(KB24_list)
            KB_list.append(KB34_list)
            KB_list.append(KB35_list)
            KB_list.append(KB36_list)
            KB_list.append(KB37_list)
            KB_list.append(KB38_list)
            # print(KB_list) 
            begin = 0
            for KB in KB_list:
                begin_index = 0
                for pin in KB:
                    pin = str(int(pin[2] + pin[7:9]) + 1)
                    KB_list[begin][begin_index] = pin
                    begin_index = begin_index + 1
                begin = begin + 1 
            # print(KB_list)
            KB_name = ["KB6", "KB7", "KB8", "KB24", "KB34", "KB35", "KB36", "KB37", "KB38"]
            iSite = 1         
            begin = 0    
            for KB in KB_name:
                # print(signal)
                file_java.write("signal " + KB + "{\n")
                for pin in KB_list[begin]:
                    # print(pin)
                    if iSite == 2:
                        # print(begin)
                        file_java.write("Site9 { pogo = " + pin+ "; }\n")
                        file_java.write("Site10 { pogo = " + pin+ "; }\n")
                        file_java.write("Site11 { pogo = " + pin+ "; }\n")
                        file_java.write("Site12 { pogo = " + pin+ "; }\n")
                        file_java.write("Site13 { pogo = " + pin+ "; }\n")
                        file_java.write("Site14 { pogo = " + pin+ "; }\n")
                        file_java.write("Site15 { pogo = " + pin+ "; }\n")
                        file_java.write("Site16 { pogo = " + pin+ "; }\n")
                        file_java.write("}\n")    
                        iSite = 1
                    else:    
                        # file_java.write("Site " + str(iSite) + "{ pogo = " + pin+ "; }\n")
                        file_java.write("Site1 { pogo = " + pin+ "; }\n")
                        file_java.write("Site2 { pogo = " + pin+ "; }\n")
                        file_java.write("Site3 { pogo = " + pin+ "; }\n")
                        file_java.write("Site4 { pogo = " + pin+ "; }\n")
                        file_java.write("Site5 { pogo = " + pin+ "; }\n")
                        file_java.write("Site6 { pogo = " + pin+ "; }\n")
                        file_java.write("Site7 { pogo = " + pin+ "; }\n")
                        file_java.write("Site8 { pogo = " + pin+ "; }\n")
                        iSite = iSite + 1
                begin = begin + 1        
file_java.close

