from os import listdir
from os.path import join
import re
from openpyxl import Workbook
import csv
from natsort import natsorted
import os
from openpyxl import load_workbook

def gen_pathList(path, flag): 
  if flag == "csv":
    pattern_csv = r'\w*csv\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      # 產生檔案的絕對路徑
        fullpath = join(path, f)
      # 判斷 fullpath 是檔案還是目錄
        if re.search(pattern_csv, fullpath) != None:
            path_list.append(fullpath)
    # print(natsorted(path_list))
    return natsorted(path_list)
  elif flag == "xlsx":
    pattern_xlsx = r'\w*xlsx\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      # 產生檔案的絕對路徑
        fullpath = join(path, f)
      # 判斷 fullpath 是檔案還是目錄
        if re.search(pattern_xlsx, fullpath) != None:
            path_list.append(fullpath)
    print(natsorted(path_list))
    return natsorted(path_list)
  elif flag == "txt":
    pattern_xlsx = r'\w*txt\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      # 產生檔案的絕對路徑
        fullpath = join(path, f)
      # 判斷 fullpath 是檔案還是目錄
        if re.search(pattern_xlsx, fullpath) != None:
            path_list.append(fullpath)
    print(natsorted(path_list))
    return natsorted(path_list)  
  elif flag == "stdf":
    pattern_stdf = r'\w*stdf\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      # 產生檔案的絕對路徑
        fullpath = join(path, f)
      # 判斷 fullpath 是檔案還是目錄
        if re.search(pattern_stdf, fullpath) != None:
            path_list.append(fullpath)
    print(natsorted(path_list))
    return natsorted(path_list) 
  elif flag == "s2p":
    pattern_stdf = r'\w*s2p\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      # 產生檔案的絕對路徑
        fullpath = join(path, f)
      # 判斷 fullpath 是檔案還是目錄
        if re.search(pattern_stdf, fullpath) != None:
            path_list.append(fullpath)
    print(natsorted(path_list))
    return natsorted(path_list)   
# gen_pathList("C:\\Users\\Brian.tsai\\Desktop\\project\\MT6177M\\datalog\\20201204_test")

def csvToexcel(path_list):
  wb_csv = Workbook()
  ws_csv = wb_csv.active
  for file in path_list:
    with open(file, 'r') as f:
      for row in csv.reader(f):
        ws_csv.append(row)
    wb_csv.save(file.replace("csv", "xlsx"))
    wb_csv.close()

def gen_bcal_file(file, golden_start):

  bcal_name = "BoardCal_forJason.csv" # 原始bcal file
  bcal_name_final = "BoardCal_forJason_final.csv" # 最後要給出去的bcal file


  pattern_testNumber_bcal = r'\w*TestNumber\w*'
  bcal_list = []  
  dict_Bcal = dataDictionary (file, 20) # test number, data 鍵值對
  with open("C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\ori_bcal\\" + bcal_name, newline='') as csvfile_bcal: # 打開原始bcal file
    test_number_index = 0
    test_number_bcal_list = []
    rows = csv.reader(csvfile_bcal)
    index = 0
    for row in rows:        
      for i in range (0, len(row), 1):
        if re.search(pattern_testNumber_bcal, row[i]) != None: 
          test_number_index = i # 找出test number 是在第幾欄           
      test_number_bcal_list.append(row[test_number_index]) # 將 "TestNumber"和test number append進去list, index數必須和原始bcal file row數一樣
      try:
        if index == 0:
          row.append("Gold" + str(golden_start)) # 將算出來的平均data append在原始bcal file後面
          print("Gold: " +  str(golden_start) ) # 印出目前是第幾個golden sample, 做記號用
          # print(row)
          bcal_list.append(row)
          index = index +1
        else:
          row.append(str(dict_Bcal[test_number_bcal_list[index]])) # 將算出來的平均data append在原始bcal file後面
          bcal_list.append(row)
          index = index +1
      except KeyError:  
        bcal_list.append(row) # 發生KeyError(因為test_number_bcal_list裡有"")
        index = index +1    
    # print(bcal_list)
    # print(test_number_bcal_list)    
  with open("C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\ori_bcal\\" + bcal_name, 'w', newline='') as csvfile_bcal_write:    
    writer = csv.writer(csvfile_bcal_write)
    for row in bcal_list:
      # print(row)
      writer.writerow(row) 
  with open("C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\ori_bcal\\" + bcal_name_final, 'w', newline='') as csvfile_bcal_write_final:    
    writer = csv.writer(csvfile_bcal_write_final)
    for row in bcal_list:
      # print(row)
      writer.writerow(row)         
       
def dataDictionary (file, average):
  pattern_testNumber = r'\w*Tests#\w*'
  pattern_PID = r'\w*PID\w*'
  dict_Bcal = {}
  index = 0
  testNumber_list = [] # 全部test number
  data_list = []
  data_list_buffer = [] # 每一列IC(不同PID)的raw data
  data_list_mean = [] # 全部IC的data之平均
  with open(file, newline='') as csvfile:
    rows = csv.reader(csvfile)
    for row in rows:
      try:
        if re.search(pattern_testNumber, row[0]) != None: 
          testNumber_list = row[10:]
          del testNumber_list[-1]
          # print(testNumber_list)
          # print(len(testNumber_list))
          # print("testNumber_list_length: " + str(len(testNumber_list)))
          data_list = [0] * len(testNumber_list)
          data_list_mean = [0] * len(testNumber_list)
        elif re.search(pattern_PID, row[0]) != None:
          # print("testNumber_list_length: " + str(len(testNumber_list)))
          data_list_buffer = row[10:]
          for i in range (0, len(testNumber_list), 1):
            # print(i)
            # print(data_list_buffer[i])
            # print(float(data_list_buffer[i]))
            try:
              data_list[i] = (data_list[i]) + float(data_list_buffer[i])  
            except ValueError:
              # print(str(i) + ": ValueError")
              continue   
          for i in range (0, len(testNumber_list), 1):
            data_list_mean[i] = data_list[i]/average
          for number in testNumber_list:
            # print(index)
            dict_Bcal[number] = data_list_mean[index] # 做出test number, data 鍵值對
            index = index + 1
          index = 0  
      except IndexError:
        continue
  # print(dict_Bcal.items())  
  return dict_Bcal

def gen_fileList(path, flag): 
  if flag == "csv":
    pattern_csv = r'\w*csv\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      if re.search(pattern_csv, f) != None:
          path_list.append(f)
    print(path_list)
    return path_list
  if flag == "xlsx":
    pattern_xlsx = r'\w*xlsx\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      if re.search(pattern_xlsx, f) != None:
          path_list.append(f)
    print(path_list)
    return path_list
  if flag == "stdf":
    pattern_stdf = r'\w*stdf\w*'
    path_list = []
    files = listdir(path)
    for f in files:
      if re.search(pattern_stdf, f) != None:
          path_list.append(f)
    print(path_list)
    return path_list  

def cal_rawdata (file):
  pattern_Gold = r'\w*Gold\w*'
  pattern_site = r'^[A-Z]*\d+\W*'
  index = 0
  site_list = []
  gold_dict = {}
  new_row = []
  header_row = []
  choose_gold =""
  with open(file, newline='') as csvfile:
    rows = csv.reader(csvfile)
    for row in rows:
      if index == 0:
        for i in range (0, len(row), 1):
          if re.search(pattern_site, row[i]) != None:
            site_list.append(i)
            # site_dict[row[i]] = i 
          elif re.search(pattern_Gold, row[i]) != None:
            gold_dict[row[i]] = i  
        print(site_list)  
        print(gold_dict)
        choose_gold = str(input("請選擇golden: "))
        header_row = row
        for site in site_list:
          header_row.append(choose_gold + "-" + row[site])
        index = index + 1
        continue
      for site in site_list:
        if row[int(gold_dict[choose_gold])] != "":
          try:
            row.append(float(row[int(gold_dict[choose_gold])]) - float(row[site]))
          except ValueError:
            continue  
      new_row.append(row)
      # print(new_row)            
  with open(file.replace(".csv", "_rawdata.csv") , 'a+', newline='') as csvfile_W:    
        writer = csv.writer(csvfile_W)
        writer.writerow(header_row)
        for row in new_row:
            writer.writerow(row) 
def gencontext_list(file):
  context_list = []
  txt_file = open(file, 'r')
  context = txt_file.readlines()
  for i in range (0, len(context), 1):
    context_list.append(str(i+1) + ". " + context[i])
  print(context_list[13])
  return context_list       

def excel_find_column(file, sheet, row_start, column_start, row_interval = 1, printOut = True): # 把某欄存到 list裡面
  wb = load_workbook(file)
  ws = wb["%s"] %sheet
  list = []
  for i in range(row_start,ws.max_row-1, row_interval):
    Ce = (ws.cell(row = i, column=column_start))
    value = Ce.value
    list.append(value)
  if printOut == True:
    print(list)    
  return list  

def unmerged_excel(src_file, sheet):
    pattern_digit= re.compile(r'\d+')
    pattern_Ndigit = re.compile(r'\D+')
    # src_file = "D:\\Project\\ProjectC\\Schematic\\final_version\\(1)TMNZ16X8_WSRF_R8_FT-V1__D01_ch_map_p.xlsx"
    wb = load_workbook(src_file)
    ws_Relay_MAP = wb[sheet]
    merge_list = ws_Relay_MAP.merged_cells.ranges # fine merge cells
    print(merge_list)
    # ws_Relay_MAP.unmerge_cells()
    while len(merge_list) != 0:
        for item in merge_list: # E5:E7
            print(item)
            item_split_1 = str(item).split(":")[0]  # E5
            item_split_2 = str(item).split(":")[1]  # E7
            ws_Relay_MAP.unmerge_cells(item_split_1 + ":" + item_split_2)  # unmege all cell
            digit_1 = int (pattern_digit.findall(item_split_1)[0]) # 5
            digit_2 = int (pattern_digit.findall(item_split_2)[0]) # 7
            Ndigit_1 = pattern_Ndigit.findall(item_split_1)[0]
            merge_range = digit_2 - digit_1
            cell = ws_Relay_MAP[item_split_1].value # value of E5
            for i in range (0, merge_range,1):
                ws_Relay_MAP[Ndigit_1+str(digit_1+i+1)] = cell
            merge_list = ws_Relay_MAP.merged_cells.ranges    
    wb.save("D:\\Project\\ProjectC\\Schematic\\final_version\\(1)TMNZ16X8_WSRF_R8_FT-V1__D01_ch_map_p_unmerged.xlsx") 
