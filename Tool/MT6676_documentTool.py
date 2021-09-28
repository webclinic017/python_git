from openpyxl import load_workbook
import re
import os
from tool_function import gen_pathList


PPG_gen = "C:\\Users\\Brian.tsai\\Desktop\\Python\\MT6671\\PPG\\"
PLL_gen = "C:\\Users\\Brian.tsai\\Desktop\\Python\\MT6671\\PLL\\"
LED_gen = "C:\\Users\\Brian.tsai\\Desktop\\Python\\MT6671\\LED\\"

PPG_doc = "D:\\Project\\MT6676\\A11167_PPGFE_FT_DOC_20210512.xlsx"
PLL_doc = "D:\\Project\\MT6676\\MT6676_PLLOSC_FT_DOC_20210510.xlsx"
LED_doc = "D:\\Project\\MT6676\\MT6676_LEDDRV_FT_DOC_20210419.xlsx"

path_list = gen_pathList(PPG_gen,"txt")
for file in path_list:
    try:
        os.remove(file)
    except OSError as e:
        print(e)
    else:
        print("File is deleted successfully")

path_list = gen_pathList(PLL_gen,"txt")
for file in path_list:
    try:
        os.remove(file)
    except OSError as e:
        print(e)
    else:
        print("File is deleted successfully")   

path_list = gen_pathList(LED_gen,"txt")
for file in path_list:
    try:
        os.remove(file)
    except OSError as e:
        print(e)
    else:
        print("File is deleted successfully")   

try:
    wb = load_workbook(PPG_doc, data_only = True)
    # wb_PLL = load_workbook("D:\\Project\\MT6676\\MT6676_PLLOSC_FT_DOC_20210510.xlsx", data_only = True)
    # wb_LED = load_workbook("D:\\Project\\MT6676\\MT6676_LEDDRV_FT_DOC_20210419.xlsx", data_only = True) 
except FileNotFoundError as e:
    print(e) 
    print("PPG file is not generated")
else:
    ws = wb.sheetnames
    # print(ws)
    pattern_IP_register = r'\w*\W*IP register\W*\w*'
    pattern_Pre_Condition = r'\w*\W*Condition\W*\w*'
    pattern_I2CW = r'\w*I2CW\W*\w*\W*\w*'
    pattern_mdelay = r'\w*mdelay\W*\w*\W*'



    for i in range (1,14):
        IP_register_row = 0
        Pre_Condition_row = 0
        workSheet = wb[ws[i]]
        workSheetName = ws[i]
        print(workSheetName)
        for i in range (1, workSheet.max_row, 1):
            Ce = workSheet.cell(row = i, column=2).value
            # print(Ce)
            if type(Ce) == str:
                if re.search(pattern_IP_register, Ce) != None:
                    # print(Ce)
                    IP_register_row = i
                elif re.search(pattern_Pre_Condition, Ce) != None: 
                    # print(Ce)
                    Pre_Condition_row = i
        # print(IP_register_row)   
        # print(Pre_Condition_row)
        rg_list = []
        for i in range (IP_register_row, Pre_Condition_row, 1):   
            Ce = workSheet.cell(row = i, column=3).value
            Ce_font = workSheet.cell(row = i, column=3)
            if type(Ce) == str:
                # print(Ce_font.font.strike)
                if Ce_font.font.strike == True:
                    try: 
                        comment = workSheet.cell(row = i, column=4).value
                        row_str = "//" + Ce + " This had been deleted" + "\t\t\t" + comment
                    except TypeError:
                        row_str = "// This had been deleted"  
                    rg_list.append(row_str)
                elif re.search(pattern_mdelay, Ce) != None:
                    mdelay = "msWait(" + Ce[7:] + ";"
                    # print(mdelay)
                    rg_list.append(mdelay)
                elif re.search(pattern_I2CW, Ce) != None:
                    if len(Ce.split(" ")[1]) < 6:
                        rg_oneByte = Ce[0:5] + "Sub.oneByte, " + Ce[5:16] + ";"
                        try: 
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_oneByte + "\t\t\t" + comment
                        except TypeError:
                            row_str = rg_oneByte     
                        # print(rg_oneByte)
                        rg_list.append(row_str)
                    else:
                        rg_fourByte = Ce[0:5] + "Sub.fourBytes, " + Ce[5:22] + ";"
                        try:
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_fourByte + "\t" + comment
                        except TypeError:
                            row_str = rg_fourByte      
                        # print(rg_fourByte)             
                        rg_list.append(row_str)
                else:
                    comment = "//" + Ce
                    rg_list.append(comment)         
        # print(rg_list)                
        f = open(PPG_gen + workSheetName + ".txt","w")
        for rg in rg_list:
            f.write(rg)
            f.write("\n")
        f.close() 

try:
    wb_PLL = load_workbook(PLL_doc, data_only = True)
except FileNotFoundError as e:
    print(e) 
    print("PLL file is not generated")
else:
    ws_PLL = wb_PLL.sheetnames
    # print(ws_PLL)
    pattern_BPA_delay_ms = r'\w*BPA_delay_ms\W*\w*\W*'
    pattern_BPA_reg_write = r'\w*BPA_reg_write\W*\w*\W*\w*\W*'

    for i in range (1,11):
        IP_register_row = 0
        Pre_Condition_row = 0
        workSheet = wb_PLL[ws_PLL[i]]
        workSheetName = ws_PLL[i]
        print(workSheetName)
        for i in range (1, workSheet.max_row, 1):
            Ce = workSheet.cell(row = i, column=2).value
            # print(Ce)
            if type(Ce) == str:
                if re.search(pattern_IP_register, Ce) != None:
                    # print(Ce)
                    IP_register_row = i
                elif re.search(pattern_Pre_Condition, Ce) != None: 
                    # print(Ce)
                    Pre_Condition_row = i
        # print(IP_register_row)   
        # print(Pre_Condition_row)
        rg_list = []
        
        for i in range (IP_register_row, Pre_Condition_row, 1):   
            Ce = workSheet.cell(row = i, column=3).value
            Ce_font = workSheet.cell(row = i, column=3)
            if type(Ce) == str:
                # print(Ce_font.font.strike)
                if Ce_font.font.strike == True:
                    try: 
                        comment = workSheet.cell(row = i, column=4).value
                        row_str = "//" + Ce + " This had been deleted" + "\t\t\t" + comment
                    except TypeError:
                        row_str = "// This had been deleted"  
                    rg_list.append(row_str)
                elif re.search(pattern_BPA_delay_ms, Ce) != None:
                    mdelay = "msWait(" + Ce[14:] + ";"
                    # print(mdelay)
                    rg_list.append(mdelay)
                elif re.search(pattern_BPA_reg_write, Ce) != None:
                    # print(Ce.split(",")[1].split(" ")[1])
                    # print(len((Ce.split(",")[1]).split(" ")[1]))
                    if len((Ce.split(",")[1]).split(" ")[1]) < 6:
                        rg_oneByte = "I2CW(" + "Sub.oneByte, " + Ce.split("(")[1][0:11] + ";"
                        try: 
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_oneByte + "\t\t\t" + comment
                        except TypeError:
                            row_str = rg_oneByte     
                        # print(rg_oneByte)
                        rg_list.append(row_str)
                    else:
                        rg_fourBytes = "I2CW(" + "Sub.fourBytes, " + Ce.split("(")[1][0:17] + ";"
                        try: 
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_fourBytes + "\t\t\t" + comment
                        except TypeError:
                            row_str = rg_fourBytes     
                        # print(rg_oneByte)
                        rg_list.append(row_str)  
                else:
                    comment = "//" + Ce
                    rg_list.append(comment)          
        f = open(PLL_gen + workSheetName + ".txt","w")
        for rg in rg_list:
            f.write(rg)
            f.write("\n")
        f.close()      

try:
    wb_LED = load_workbook(LED_doc, data_only = True) 
except FileNotFoundError as e:
    print(e) 
    print("LED file is not generated")
else:
    pattern_BPA_delay_ms = r'\w*BPA_delay_ms\W*\w*\W*'
    pattern_BPA_reg_write = r'\w*BPA_reg_write\W*\w*\W*\w*\W*'
    ws_LED = wb_LED.sheetnames
    # print(ws_LED)  
    for i in range (1,5):
        IP_register_row = 0
        Pre_Condition_row = 0
        workSheet = wb_LED[ws_LED[i]]
        workSheetName = ws_LED[i]
        print(workSheetName)
        for i in range (1, workSheet.max_row, 1):
            Ce = workSheet.cell(row = i, column=2).value
            # print(Ce)
            if type(Ce) == str:
                if re.search(pattern_IP_register, Ce) != None:
                    # print(Ce)
                    IP_register_row = i
                elif re.search(pattern_Pre_Condition, Ce) != None: 
                    # print(Ce)
                    Pre_Condition_row = i
        # print(IP_register_row)   
        # print(Pre_Condition_row)
        rg_list = []
        
        for i in range (IP_register_row, Pre_Condition_row, 1):   
            Ce = workSheet.cell(row = i, column=3).value
            Ce_font = workSheet.cell(row = i, column=3)
            if type(Ce) == str:
                # print(Ce_font.font.strike)
                if Ce_font.font.strike == True:
                    try: 
                        comment = workSheet.cell(row = i, column=4).value
                        row_str = "//" + Ce + " This had been deleted" + "\t\t\t" + comment
                    except TypeError:
                        row_str = "// This had been deleted"  
                    rg_list.append(row_str)
                elif re.search(pattern_BPA_delay_ms, Ce) != None:
                    mdelay = "msWait(" + Ce[14:] + ";"
                    # print(mdelay)
                    rg_list.append(mdelay)
                elif re.search(pattern_BPA_reg_write, Ce) != None:
                    # print(Ce.split(",")[1].split(" ")[1])
                    # print(len((Ce.split(",")[1]).split(" ")[1]))
                    if len((Ce.split(",")[1]).split(" ")[1]) < 6:
                        rg_oneByte = "I2CW(" + "Sub.oneByte, " + Ce.split("(")[1][0:11] + ";"
                        try: 
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_oneByte + "\t\t\t" + comment
                        except TypeError:
                            row_str = rg_oneByte     
                        # print(rg_oneByte)
                        rg_list.append(row_str)
                    else:
                        rg_fourBytes = "I2CW(" + "Sub.fourBytes, " + Ce.split("(")[1][0:17] + ";"
                        try: 
                            comment = "//" + workSheet.cell(row = i, column=4).value
                            row_str = rg_fourBytes + "\t\t\t" + comment
                        except TypeError:
                            row_str = rg_fourBytes     
                        # print(rg_oneByte)
                        rg_list.append(row_str)   
                else:
                    comment = "//" + Ce
                    rg_list.append(comment)         
        f = open(LED_gen + workSheetName + ".txt","w")
        for rg in rg_list:
            f.write(rg)
            f.write("\n")
        f.close()       



