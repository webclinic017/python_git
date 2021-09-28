from openpyxl import load_workbook
import os
import re
import time

start = time.time()
try:
    os.remove("D:\\Project\\ProjectH\\Schematic\\Dutboard_ChannelMap_1_site_FT.dbd")
except OSError as e:
    print(e)
else:
    print("================================File is deleted successfully================================")

pattern_DGND = r'DGND'
pattern_5 = r'\d{5}' 
pattern_power_2 = r'\d{3}_P\d{2}' 
pattern_power_1 = r'\d{3}_P\d{1}'
power_pin_list = []

file_dbd = open("D:\\Project\\ProjectH\\Schematic\\Dutboard_ChannelMap_1_site_FT.dbd",'w') 
file_dbd.write("dutboard Dutboard_ChannelMap_1_site_FT\n")
file_dbd.write("{\n")
file_dbd.write("\tsites = 1;\n")
file = "D:\\Project\\ProjectH\\Schematic\\(1)TMPZ78_LVX_WSRF_FT-V1~~D01_FinalPA.xlsx"
wb = load_workbook(file)

ws_PA = wb['BGA']

# content = ws_PA.cell(row=1, column=10).value
# print(type(content))
for i in range (1, ws_PA.max_row+1, 1):    
    content = ws_PA.cell(row=i, column=7).value  # content = pogo
    if re.search(pattern_5, str(content)) != None:  
        # print(content)
        pin_name = str(ws_PA.cell(row=i, column=4).value)
        file_dbd.write("\tsignal " + pin_name + "\n")
        file_dbd.write("\t\t{\n")
        file_dbd.write("\t\t\tsite " + str(1) + "\n" ) 
        file_dbd.write("\t\t\t{\n")
        file_dbd.write("\t\t\t\tpogo = "  + str(content) + ";\n")
        file_dbd.write("\t\t\t}\n")    
        file_dbd.write("\t\t}\n")  
        # continue
    elif re.search(pattern_power_1, str(content)) != None:
        # print(str(ws_PA.cell(row=i, column=4).value))
        
        if content in power_pin_list:
            continue
        else:
            power_pin_list.append(content)     
            content = content.split(",")    
            if len(content) == 1:
                if re.search(pattern_power_2, str(content[0])) != None:
                    pin_name = str(ws_PA.cell(row=i, column=4).value)
                    pogo = content[0].replace("_P","")
                    file_dbd.write("\tsignal " + pin_name + "\n")
                    file_dbd.write("\t\t{\n")
                    file_dbd.write("\t\t\tsite " + str(1) + "\n" ) 
                    file_dbd.write("\t\t\t{\n")
                    file_dbd.write("\t\t\t\tpogo = "  + str(pogo) + ";\n")
                    file_dbd.write("\t\t\t}\n")    
                    file_dbd.write("\t\t}\n")  
                else:    
                    pin_name = str(ws_PA.cell(row=i, column=4).value)
                    pogo = content[0].replace("_","").replace("P","0")
                    file_dbd.write("\tsignal " + pin_name + "\n")
                    file_dbd.write("\t\t{\n")
                    file_dbd.write("\t\t\tsite " + str(1) + "\n" ) 
                    file_dbd.write("\t\t\t{\n")
                    file_dbd.write("\t\t\t\tpogo = "  + str(pogo) + ";\n")
                    file_dbd.write("\t\t\t}\n")    
                    file_dbd.write("\t\t}\n")  
            elif len(content) == 2:
                pogo_list = []
                for j in range (0, 2, 1):
                    if re.search(pattern_power_2, str(content[j])) != None:
                        pogo = content[j].replace("_P","")
                        pogo_list.append(pogo)
                    else:
                        pogo = content[j].replace("_","").replace("P","0")
                        pogo_list.append(pogo)
                pin_name = str(ws_PA.cell(row=i, column=4).value)   
                # print(pin_name)     
                file_dbd.write("\tsignal " + pin_name + "\n")
                file_dbd.write("\t\t{\n")
                file_dbd.write("\t\t\tsite " + str(1) + "\n" ) 
                file_dbd.write("\t\t\t{\n")
                file_dbd.write("\t\t\t\tpogo = "  + pogo_list[0] + "|" + pogo_list[1] +  ";\n")
                file_dbd.write("\t\t\t}\n")    
                file_dbd.write("\t\t}\n")  
            elif len(content) == 3:    
                pogo_list = []
                for j in range (0, 3, 1):
                    if re.search(pattern_power_2, str(content[j])) != None:
                        pogo = content[j].replace("_P","")
                        pogo_list.append(pogo)
                    else:
                        pogo = content[j].replace("_","").replace("P","0")
                        pogo_list.append(pogo)
                pin_name = str(ws_PA.cell(row=i, column=4).value)        
                file_dbd.write("\tsignal " + pin_name + "\n")
                file_dbd.write("\t\t{\n")
                file_dbd.write("\t\t\tsite " + str(1) + "\n" ) 
                file_dbd.write("\t\t\t{\n")
                file_dbd.write("\t\t\t\tpogo = "  + pogo_list[0] + "|" + pogo_list[1] +  "|" + pogo_list[2] + ";\n")
                file_dbd.write("\t\t\t}\n")    
                file_dbd.write("\t\t}\n")          
# print(power_pin_list)        

file_dbd.write("}\n") 
file_dbd.close
end = time.time()
print("program time: " + str(end-start) + "s")