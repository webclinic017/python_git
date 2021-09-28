from tool_function import unmerged_excel
import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook


src_file = "D:\\Project\\ProjectC\\Schematic\\final_version\\(1)TMNZ16X8_WSRF_R8_FT-V1__D01_ch_map_p.xlsx"
gen_file = "D:\\Project\\ProjectC\\Schematic\\final_version\\ChannelMap_8_site_FT.xlsx"
try:
    os.remove(gen_file)
except OSError as e:
    print(e)
else:
    print("================================File is deleted successfully================================")


def remove_char(string):
    newstring = re.sub(r'[A-Z]+', '', string)
    return newstring


pattern_Utility = r'\d+U\w+'
pattern_slash = r'\d+\W{1}\d+'
pattern_VDD =  r'VDD_\W*'
pattern_SMP =  r'\d*SMP\d*'
pattern_pogo = r'\d{5}'
pattern_426 = r'426_\d+'
pattrn_FUSE = r'FUSE_\w*'
pattrn_K = r'\d+K\d+'
pattern_CH = r'CH_'
pattern_Relay = r'Relay'

RF_mapping = {"1SMP1":"90201", "2SMP1":"90101", "3SMP1":"90901", "4SMP1":"91001", "5SMP1":"90209", "6SMP1":"90109", "7SMP1":"90909", "8SMP1":"91009",
              "1SMP4":"90202", "2SMP4":"90102", "3SMP4":"90902", "4SMP4":"91002", "5SMP4":"90210", "6SMP4":"90110", "7SMP4":"90910", "8SMP4":"91010",
              "1SMP11":"90203", "2SMP11":"90103", "3SMP11":"90903", "4SMP11":"91003", "5SMP11":"90211", "6SMP11":"90111", "7SMP11":"90911", "8SMP11":"91011",
              "1SMP13":"90205", "2SMP13":"90105", "3SMP13":"90905", "4SMP13":"91005", "5SMP13":"90213", "6SMP13":"90113", "7SMP13":"90913", "8SMP13":"91013",
              "1SMP14":"90206", "2SMP14":"90106", "3SMP14":"90906", "4SMP14":"91006", "5SMP14":"90214", "6SMP14":"90114", "7SMP14":"90914", "8SMP14":"91014",
              "1SMP9":"90207", "2SMP9":"90107", "3SMP9":"90907", "4SMP9":"91007", "5SMP9":"90215", "6SMP9":"90115", "7SMP9":"90915", "8SMP9":"91015"}

wb_gen = Workbook()
ws_gen = wb_gen["Sheet"]
ws_gen.title = "ChannelMap_8_site_FT"    

ws_gen.cell(row=1, column=1, value="Net Name")
ws_gen.cell(row=1, column=2, value="Type")
ws_gen.cell(row=1, column=3, value="Site1")
ws_gen.cell(row=1, column=4, value="Site2")
ws_gen.cell(row=1, column=5, value="Site3")
ws_gen.cell(row=1, column=6, value="Site4")
ws_gen.cell(row=1, column=7, value="Site5")
ws_gen.cell(row=1, column=8, value="Site6")
ws_gen.cell(row=1, column=9, value="Site7")
ws_gen.cell(row=1, column=10, value="Site8")

sheet_Map = {}
site_dict = {}
for i in range (1, ws_gen.max_column+1, 1):
    content = ws_gen.cell(row=1, column=i).value
    site_dict[content] = i
print(site_dict)    

wb = load_workbook(src_file)
# print(wb.sheetnames)
for sheet in wb.sheetnames:
    if re.search(pattern_CH, sheet) != None:
        sheet_Map["Channel"] = sheet
    elif re.search(pattern_Relay, sheet) != None:
        sheet_Map["Relay"] = sheet    
print(sheet_Map)        

unmerged_excel(src_file, sheet_Map["Relay"])

wb = load_workbook(src_file)
ws_CH_MAP = wb[sheet_Map["Channel"]]
ws_Relay_MAP = wb[sheet_Map["Relay"]]

Net_name = {}
for i in range (2, ws_CH_MAP.max_row+1, 1):
    content = ws_CH_MAP.cell(row=i, column=2).value
    if type(content) == str:
        if re.search(pattern_VDD, content) != None:
            pin_name = content
            if pin_name in Net_name:
                    continue
            else:
                # print(pin_name)
                Net_name[pin_name] = ws_gen.max_row + 1
                ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                ws_gen.cell(row=ws_gen.max_row, column=2, value="POWER")
        elif re.search(pattrn_FUSE, content) != None:
            pin_name = content
            if pin_name in Net_name:
                    continue
            else:
                # print(pin_name)
                Net_name[pin_name] = ws_gen.max_row + 1
                ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)     
                ws_gen.cell(row=ws_gen.max_row, column=2, value="POWER")   
        else:
            pogo = ws_CH_MAP.cell(row=i, column=3).value
            if pogo == "DGND":
                continue
            elif len(pogo.split(",")) == 2:
                pin_name_static = content
                pin_name = content
                for index in range (0, 2, 1):
                    if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                        pin_name = pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    elif re.search(pattern_SMP, pogo.split(",")[index]) != None:  
                        pin_name = "RF_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                        pin_name = "MX_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    else:
                        pin_name = pin_name_static + "_" + pogo.split(",")[index]    
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
            elif len(pogo.split(",")) == 3:
                pin_name_static = content
                pin_name = content
                for index in range (0, 3, 1):
                    if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                        pin_name = pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    elif re.search(pattern_SMP, pogo.split(",")[index]) != None: 
                        pin_name = "RF_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name) 
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")           
                    elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                        pin_name = "MX_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)  
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    else:
                        pin_name = pin_name_static + "_" +  pogo.split(",")[index]    
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
            elif len(pogo.split(",")) == 4:  
                pin_name_static = content
                pin_name = content
                for index in range (0, 4, 1):
                    if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                        pin_name = pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    elif re.search(pattern_SMP, pogo.split(",")[index]) != None: 
                        pin_name = "RF_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)   
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")         
                    elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                        pin_name = "MX_" + pin_name_static + "_" + str(index)
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)  
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   
                    else:
                        pin_name = pin_name_static + "_" +  pogo.split(",")[index]    
                        Net_name[pin_name] = ws_gen.max_row + 1
                        ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                        ws_gen.cell(row=ws_gen.max_row, column=2, value="Channel")   

for i in range (2, ws_CH_MAP.max_row+1, 1):
    content = ws_CH_MAP.cell(row=i, column=2).value
    if type(content) == str:
        if re.search(pattern_VDD, content) != None:
            pin_name = content
            for j in range(3, ws_CH_MAP.max_column, 1):
                # print(ws_CH_MAP.max_column)
                site = "Site" + str(j-2)
                pogo_num = ws_CH_MAP.cell(row=i, column=j).value
                # print(pogo_num)
                if len(pogo_num.split(",")) == 2:
                    pogo_num_split = pogo_num.split(",")
                    pogo_num = pogo_num_split[0].split("_")[0] + "-" + pogo_num_split[0].split("_")[1] + "-" + pogo_num_split[1].split("_")[1]  
                    if pin_name in Net_name:
                        ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo_num)
                else:
                    pogo_num = pogo_num.split("_")[0] + "-" + pogo_num.split("_")[1]
                    if pin_name in Net_name:
                        ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo_num)
        elif re.search(pattrn_FUSE, content) != None:
            pin_name = content
            for j in range(3, ws_CH_MAP.max_column, 1):
                site = "Site" + str(j-2)
                pogo_num = ws_CH_MAP.cell(row=i, column=j).value
                pogo_num = pogo_num.split("_")[0] + "-" + pogo_num.split("_")[1]
                if pin_name in Net_name:
                    ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo_num)
        else:
            for j in range(3, ws_CH_MAP.max_column, 1):
                pogo = str(ws_CH_MAP.cell(row=i, column=j).value)
                if pogo == "DGND":
                    continue
                elif len(pogo.split(",")) == 2:
                    pin_name_static = content
                    pin_name = content                   
                    site = "Site" + str(j-2)
                    for index in range (0, 2, 1):
                        if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                            pin_name = pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])
                        elif re.search(pattern_SMP, pogo.split(",")[index]) != None:
                            pin_name = "RF_" + pin_name_static + "_" + str(index)
                            try:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=RF_mapping[pogo.split(",")[index]])
                            except KeyError:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])        
                        elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                            pin_name = "MX_" + pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])      
                        else:
                            try:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]    
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])   
                            except KeyError:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]
                                Net_name[pin_name] = ws_gen.max_row + 1
                                ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])   
                elif len(pogo.split(",")) == 3:
                    pin_name_static = content
                    pin_name = content                   
                    site = "Site" + str(j-2)
                    for index in range (0, 3, 1):
                        if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                            pin_name = pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])
                        elif re.search(pattern_SMP, pogo.split(",")[index]) != None:
                            pin_name = "RF_" + pin_name_static + "_" + str(index)
                            try:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=RF_mapping[pogo.split(",")[index]])
                            except KeyError:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])        
                        elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                            pin_name = "MX_" + pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])      
                        else:
                            try:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]    
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])   
                            except KeyError:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]
                                Net_name[pin_name] = ws_gen.max_row + 1
                                ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])  
                elif len(pogo.split(",")) == 4:
                    pin_name_static = content
                    pin_name = content                   
                    site = "Site" + str(j-2)
                    for index in range (0, 4, 1):
                        if re.search(pattern_pogo, pogo.split(",")[index]) != None:
                            pin_name = pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])
                        elif re.search(pattern_SMP, pogo.split(",")[index]) != None:
                            pin_name = "RF_" + pin_name_static + "_" + str(index)
                            try:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=RF_mapping[pogo.split(",")[index]])
                            except KeyError:
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])        
                        elif re.search(pattern_426, pogo.split(",")[index]) != None: 
                            pin_name = "MX_" + pin_name_static + "_" + str(index)
                            ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])      
                        else:
                            try:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]    
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])   
                            except KeyError:
                                pin_name = pin_name_static + "_" +  pogo.split(",")[index]
                                Net_name[pin_name] = ws_gen.max_row + 1
                                ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                                ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=pogo.split(",")[index])       

for i in range (1, ws_Relay_MAP.max_column+1, 1):
    for j in range (1, ws_Relay_MAP.max_row+1, 1):
        content = ws_Relay_MAP.cell(row=j, column=i).value
        if type(content) == str:
            if re.search(pattern_slash, content) != None: 
                content_split = content.split("/")
                print(content)
                site = "Site" + content_split[0][0]
                pin_name = content_split[0][1:] + "_" + ws_Relay_MAP.cell(row=j, column=i+1).value
                if pin_name in Net_name:
                    continue
                else:
                    Net_name[pin_name] = ws_gen.max_row + 1
                    ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                    ws_gen.cell(row=ws_gen.max_row, column=2, value="Relay")
                print(site)
                print(pin_name)
            elif re.search(pattern_Utility, content) != None: 
                pin_name = content.strip()[1:] + "_" + ws_Relay_MAP.cell(row=j, column=i+1).value.strip()
                if pin_name in Net_name:
                    continue
                else:
                    Net_name[pin_name] = ws_gen.max_row + 1
                    ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                    ws_gen.cell(row=ws_gen.max_row, column=2, value="Relay")
            elif re.search(pattrn_K, content) != None:
                pin_name = content.strip()[1:]
                if pin_name in Net_name:
                    continue
                else:
                    Net_name[pin_name] = ws_gen.max_row + 1
                    ws_gen.cell(row=ws_gen.max_row + 1, column=1, value=pin_name)
                    ws_gen.cell(row=ws_gen.max_row, column=2, value="Relay")   

# print(Net_name)

for i in range (1, ws_Relay_MAP.max_column+1, 1):
    for j in range (1, ws_Relay_MAP.max_row+1, 1):
        content = ws_Relay_MAP.cell(row=j, column=i).value
        if type(content) == str:
            if re.search(pattern_slash, content) != None: 
                content_split = content.split("/")                
                site_1 = "Site" + content_split[0][0]
                site_2 = "Site" + content_split[1][0]
                pin_name = content_split[0][1:] + "_" + ws_Relay_MAP.cell(row=j, column=i+1).value
                utility_tmp = ws_Relay_MAP.cell(row=j, column=i+2).value.split("_")
                utility = "URW" + remove_char(utility_tmp[0]) + "-" + utility_tmp[1]
                if pin_name in Net_name:
                    ws_gen.cell(row=Net_name[pin_name], column=site_dict[site_1], value=utility)
                    ws_gen.cell(row=Net_name[pin_name], column=site_dict[site_2], value=utility)
            elif re.search(pattern_Utility, content) != None: 
                site = "Site" + content.strip()[0]
                pin_name = content.strip()[1:] + "_" + ws_Relay_MAP.cell(row=j, column=i+1).value.strip()
                utility_tmp = ws_Relay_MAP.cell(row=j, column=i+2).value.split("_")
                utility = "URW" + remove_char(utility_tmp[0]) + "-" + utility_tmp[1]
                # print(site)
                # print(utility)
                if pin_name in Net_name:
                    ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=utility)    
            elif re.search(pattrn_K, content) != None:
                pin_name = content.strip()[1:]        
                site = "Site" + content.strip()[0]   
                utility_tmp = ws_Relay_MAP.cell(row=j, column=i+2).value.split("_")
                utility = "URW" + remove_char(utility_tmp[0]) + "-" + utility_tmp[1] 
                if pin_name in Net_name:
                    ws_gen.cell(row=Net_name[pin_name], column=site_dict[site], value=utility)
                        
# print(Net_name)
wb_gen.save(gen_file)