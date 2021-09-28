from openpyxl import load_workbook
from openpyxl import Workbook
import re
import os
import time

start = time.time()

# path_list = gen_pathList("C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\LVC\\","xlsx")
# for file in path_list:
#     print(file)

src_file = "C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\LVC\\Carpo FTx6x2 final BOM.xlsx"
gen_file = "C:\\Users\\Brian.tsai\\Desktop\\Python\\Tool\\LVC\\Carpo FTx6x2 final BOM_mod.xlsx"

try:
    os.remove(gen_file)
except OSError as e:
    print(e)
else:
    print("================================File is deleted successfully================================")

wb_gen = Workbook()
ws_gen = wb_gen["Sheet"]
ws_gen.title = "Coverage_Report"    

ws_gen.cell(row=1, column=1, value="No.")
ws_gen.cell(row=1, column=2, value="Components")
ws_gen.cell(row=1, column=3, value="Component Type")
ws_gen.cell(row=1, column=4, value="Value")
ws_gen.cell(row=1, column=5, value="CAP Bank Name")
ws_gen.cell(row=1, column=6, value="Qty")
ws_gen.cell(row=1, column=7, value="Test Item")
ws_gen.cell(row=1, column=8, value="Test Environment")
ws_gen.cell(row=1, column=9, value="Comment")


wb = load_workbook(src_file)
pattern_BOM = r'BOM'
pattern_No = r'^No'
pattern_Reference = r'^Reference'
pattern_Description = r'Description'
pattern_Quantity = r'Quantity'
pattern_RES = r'RES'
pattern_Resistor = r'Resistor'
pattern_CAP = r'CAP'
pattern_Capacitor = r'Capacitor'
pattern_IC = r'IC'
pattern_CONNECTOR = r'CONNECTOR'
pattern_Inductor = r'Inductor'
pattern_REGULATOR = r'REGULATOR'
pattern_IC_Analog = r'IC_Analog'
pattern_IC_Logic = r'IC_Logic'
pattern_JUMPER = r'JUMPER'
pattern_Other = r'Other'
pattern_STIFFENER = r'STIFFENER'
pattern_COVER = r'COVER'

BOM_sheet = ""
column_Map = {}
column_List = []
# print(wb.sheetnames)
for sheet in wb.sheetnames:
    if re.search(pattern_BOM, sheet) != None:
        BOM_sheet = sheet
print(BOM_sheet)      

ws_BOM = wb[BOM_sheet]

for i in range (1, ws_BOM.max_row+1, 1):
    for j in range (1, ws_BOM.max_column+1, 1):
        content = ws_BOM.cell(row=i, column=j).value
        if type(content) == str:
            if re.search(pattern_No, content) != None:
                column_Map[content] = j
                column_List.append(j)
            elif re.search(pattern_Quantity, content) != None:
                column_Map[content] = j
                column_List.append(j)    
            elif  re.search(pattern_Reference, content) != None:
                column_Map[content] = j 
                column_List.append(j)   
            elif  re.search(pattern_Description, content) != None:
                column_Map[content] = j 
                column_List.append(j) 
                break
print(column_Map)            
# print(column_List)

# print(ws_BOM.max_row)
for i in range (3, ws_BOM.max_row+1, 1):
    Number = 0
    repeat_num = 0
    component = []
    Description = []
    Value = ""
    for j in column_List:
        content = ws_BOM.cell(row=i, column=j).value
        if (j == 1):
            Number = content
            if type(Number) != int:
                break
        elif (j == 5):
            repeat_num = content
        elif (j == 6):
            try:
                component = content.split(",")
                # print(component)
            except AttributeError as e:
                print("Error message: " + e)    
                print("Row: " + str(i) + ", Column: " + str(j))
        elif ( j == 8):
            if re.search(pattern_RES, content) != None:
                Description = content.split("~")
                Value = Description[4]
                Description = "Resistor"
            elif re.search(pattern_Resistor, content) != None:
                Description = content.split("~")
                Value = Description[4].split("/")[0]
                Description = "Resistor"    
            elif re.search(pattern_CAP, content) != None:  
                Description = content.split("~")
                Value = Description[5]
                Description = "Capacitor"  
            elif re.search(pattern_Capacitor, content) != None:    
                Description = content.split("~")
                Value = Description[4].split("/")[0]
                Description = "Capacitor"
            elif re.search(pattern_REGULATOR, content) != None:    
                Description = content.split("~")
                Value = Description[2]
                Description = Description[0] + "_" + Description[1]  
            elif re.search(pattern_IC_Analog, content) != None:    
                Description = content.split("~")
                Value = Description[-1]
                Description = Description[0] + "_" + Description[1]    
            elif re.search(pattern_IC_Logic, content) != None:    
                Description = content.split("~")
                Value = Description[-1]
                Description = Description[0] + "_" + Description[1]          
            elif re.search(pattern_CONNECTOR, content) != None:    
                Description = content.split("~")
                Value = Description[-1]
                Description = "Connector"  
            elif re.search(pattern_Inductor, content) != None:    
                Description = content.split("~")
                Value = Description[4]
                Description = "Inductor"    
            elif re.search(pattern_JUMPER, content) != None:    
                Description = content.split("~")
                Value = Description[4]
                Description = "JUMPER"   
            elif re.search(pattern_Other, content) != None:    
                Description = content.split("~")
                Value = Description[-1]
                Description = Description[1]           
            elif re.search(pattern_STIFFENER, content) != None:    
                Description = content.split("-")
                Value = content
                Description = Description[0]      
            elif re.search(pattern_COVER, content) != None:    
                Description = content.split("-")
                Value = content
                Description = Description[0]                     
            else:        
                try:
                    Value = content
                    Description = content.split("~")
                    Description = Description[0] + "_Not in my case"
                except AttributeError as e:
                    print("Error message: " + e)    
                    print("Row: " + str(i) + ", Column: " + str(j))     

    for repeat in range(1,repeat_num+1,1):
        ws_gen.cell(row=ws_gen.max_row+1, column=1, value=Number)
        ws_gen.cell(row=ws_gen.max_row, column=2, value=component[repeat-1])  
        ws_gen.cell(row=ws_gen.max_row, column=3, value=Description)    
        ws_gen.cell(row=ws_gen.max_row, column=4, value=Value)   
        ws_gen.cell(row=ws_gen.max_row, column=6, value=repeat_num)  
    
wb_gen.save(gen_file)

end = time.time()
print("Program time: " + str(end-start) + "(s)")