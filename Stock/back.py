import datetime
import backtrader as bt
import backtrader.feeds as btfeeds
import math
import os
import time
from stock_num_function import yoy
from openpyxl import Workbook
from openpyxl.styles import PatternFill

start = time.time()
relative_path = os.getcwd()
dt_start = datetime.datetime.strptime("20210701", "%Y%m%d")
dt_end = datetime.datetime.strptime("20210923", "%Y%m%d")
gen_file = relative_path + "\\profitDetail\\profitDetail.xlsx"

profit_dict = {}
stock_list = yoy()
print(stock_list)

# stock_list = ["2611"]
# start_cash = 100000
wb_gen = Workbook()
fill_red = PatternFill('solid', fgColor='FF0000')
fill_green = PatternFill('solid', fgColor='00FF00')
ws_gen = wb_gen["Sheet"]
ws_gen.title = "Profit_report"    
ws_gen.cell(row=1, column=1, value="Stock Number")
ws_gen.cell(row=1, column=2, value="Profit(%)")

try:
    os.remove(gen_file)
except OSError as e:
    print(e)
else:
    print("================================File is deleted successfully================================")
        
for num in stock_list:
    start_cash = 100000
    print("current stock number: " + num)
    try:
        os.remove(relative_path + "\\profitDetail\\message_" + num + ".txt")
    except OSError as e:
        print(e)
    else:
        print("================================File is deleted successfully================================")


    sourceFile = open(relative_path + "\\profitDetail\\message_" + num + ".txt", "w", encoding='utf-8')

    data = bt.feeds.GenericCSVData(
        dataname= relative_path + "\\dataFeed\\" + num + ".csv",
        fromdate=dt_start,      # 起止日期
        todate=dt_end,
        nullvalue=0.0,
        dtformat=('%Y-%m-%d'),  # 日期列的格式
        datetime=0,             # 各列的位置，从0开始，如列缺失则为None，-1表示自动根据列名判断
        open=1,
        high=2,
        low=3,
        close=4,
        volume=5,
        openinterest=-1
    )

    # sma cross strategy
    class SmaCross(bt.Strategy):
        # 交易紀錄
        def log(self, txt, dt=None):
            dt = dt or self.datas[0].datetime.date(0)
            print('%s, %s' % (dt.isoformat(), txt), file=sourceFile)
        
        # 設定交易參數
        params = dict(
            ma_period_short=5,
            ma_period_long=10
        )

        def __init__(self):
            # 均線交叉策略
            sma1 = bt.ind.SMA(period=self.p.ma_period_short)
            sma2 = bt.ind.SMA(period=self.p.ma_period_long)
            self.lines.top = bt.indicators.BollingerBands(self.datas[0], period=20).top
            self.lines.bot = bt.indicators.BollingerBands(self.datas[0], period=20).bot
            self.crossover = bt.ind.CrossOver(sma1, sma2)
            
            # 使用自訂的sizer函數，將帳上的錢all-in
            self.setsizer(sizer())
            
            # open
            self.dataopen = self.datas[0].open
            # close
            self.dataclose = self.datas[0].close
            # volume
            self.datavolume = self.datas[0].volume

        def next(self):
            print("========================================", file=sourceFile)
            print("data time: "  + str(self.datas[0].datetime.date(0)), file=sourceFile)
            # print("type data time: "  + str(type(self.datatime[0])), file=sourceFile)
            # print('當前可用資金', self.broker.getcash())
            print('當前可用資金', self.broker.getcash(), file=sourceFile)
            # print('當前總資產', self.broker.getvalue())
            print('當前總資產', self.broker.getvalue(), file=sourceFile)
            # print('當前持倉量', self.broker.getposition(self.data).size)
            print('當前持倉量', self.broker.getposition(self.data).size, file=sourceFile)
            # print('當前持倉成本', self.broker.getposition(self.data).price)
            print('當前持倉成本', self.broker.getposition(self.data).price, file=sourceFile)
            print(self.position, file=sourceFile)
            print("Open_init: " + str(self.dataopen[0]), file=sourceFile)
            print("Close_init: " + str(self.dataclose[0]), file=sourceFile)
            print("Volume_init: " + str(self.datavolume[0]), file=sourceFile)
            # 帳戶沒有部位
            if not self.position:
                # 5ma往上穿越20ma
                # print("bool top: " + str(self.lines.top[0]))
                # print("bool bot: " + str(self.lines.bot[0]))
                if self.crossover > 0 and self.dataclose[0] < 1.3 * self.lines.bot[0] : 

                    # print("close: " + str(self.dataclose[0]))
                    # print("bool: " + str(1.3 * self.lines.bot[0]))
                    # 印出買賣日期與價位
                    self.log('BUY ' + ', Price: ' + str(self.dataopen[0]))
                    # print("Volume: " + str(self.datavolume[0]))
                    # 使用開盤價買入標的
                    # self.buy(price=self.dataopen[0])
                    self.order = self.buy()

            # 5ma往下穿越20ma
            elif self.crossover < 0:
                # 印出買賣日期與價位
                self.log('SELL ' + ', Price: ' + str(self.dataopen[0]))
                # 使用開盤價賣出標的
                self.close(price=self.dataopen[0])

    # 計算交易部位
    class sizer(bt.Sizer):
        def _getsizing(self, comminfo, cash, data, isbuy):
            if isbuy:          
                # print(data[0])
                # print(data[1])
                # print(math.floor(cash/data[0]))
                # print(math.floor(cash/data[1]))
                # print("position___ : " + str(self.broker.getposition(data)))
                print("position buy : " + str(self.broker.getposition(data)), file=sourceFile)
                # print("Open : " + str(data[0]), file=sourceFile)
                print("Open buy: " + str(data.open[0]), file=sourceFile)
                # print("cash: " + str(cash))
                size = math.floor(cash/(data.open[0])) #size=股數
                size = int(size/1000) * 1000
                if size > data.volume[0]/10:
                    size = math.floor(data.volume[0]/10) 
                else:
                    size = size - 1000    
                print("Size buy: " + str(size), file=sourceFile)   
                return size
            else:
                size = self.broker.getposition(data)
                print("Size sell: " + str(size), file=sourceFile)  
                return self.broker.getposition(data)

    # 初始化cerebro
    cerebro = bt.Cerebro()
    # feed data
    cerebro.adddata(data)
    # add strategy
    cerebro.addstrategy(SmaCross)

    cerebro.broker.setcash(start_cash)
    cerebro.broker.setcommission(commission=0.002)

    # run backtest
    print("Init money: %.2f"  % start_cash, file=sourceFile)
    cerebro.run()
    final_money = cerebro.broker.getvalue()
    print("========================================", file=sourceFile)
    print('Final money: %.2f' % final_money, file=sourceFile)
    profit = 100 * (final_money - start_cash) / start_cash
    print("profit percentage: %.2f" % profit + "%", file=sourceFile)

    profit_dict[num] = profit

    sourceFile.close()
# plot diagram
    # cerebro.plot()
# print(profit_dict)

stort_row = 2
for key in profit_dict.keys():
    ws_gen.cell(row=stort_row, column=1, value=key)
    ws_gen.cell(row=stort_row, column=2, value=profit_dict[key])
    if profit_dict[key] > 10:
        ws_gen.cell(row=stort_row, column=2).fill = fill_red
    elif profit_dict[key] < -10:
        ws_gen.cell(row=stort_row, column=2).fill = fill_green
    stort_row = stort_row + 1

wb_gen.save(gen_file)
print("Generate Report")
end = time.time()
delta = end - start
print("Program time: " + str(delta))

